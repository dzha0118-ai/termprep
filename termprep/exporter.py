"""Multi-format terminology export.

Supports CSV, XLSX (Excel), TBX (ISO 30042), and JSON formats.
The TBX format is compatible with CAT tools like SDL Trados and memoQ.
"""

import json
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import Any

from termprep.db import TermDB


def export_csv(db: TermDB, output_path: str, status_filter: str | None = None) -> int:
    """Export to CSV (UTF-8 BOM for Excel compatibility)."""
    count = db.export_csv(output_path, status_filter=status_filter)
    # Add BOM for Excel to auto-detect UTF-8
    content = Path(output_path).read_bytes()
    if not content.startswith(b"\xef\xbb\xbf"):
        Path(output_path).write_bytes(b"\xef\xbb\xbf" + content)
    return count


def export_xlsx(db: TermDB, output_path: str, status_filter: str | None = None) -> int:
    """Export to Excel (.xlsx) with formatted columns and headers."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    rows = _fetch_rows(db, status_filter)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TermBase"

    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = ["Word", "Translation", "Type", "Domain", "Status", "Language", "Notes"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data rows
    for r_idx, row in enumerate(rows, 2):
        for c_idx, key in enumerate(["word", "translation", "type", "domain",
                                     "status", "lang", "notes"], 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=row.get(key, ""))
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 30

    # Auto-filter
    ws.auto_filter.ref = f"A1:G{len(rows) + 1}"

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(output_path)
    return len(rows)


def export_tbx(db: TermDB, output_path: str, status_filter: str | None = None) -> int:
    """Export to TBX (TermBase eXchange, ISO 30042) format.

    Compatible with SDL Trados, memoQ, and other CAT tools.
    """
    rows = _fetch_rows(db, status_filter)

    # Build XML
    martif = ET.Element("martif", attrib={
        "type": "TBX",
        "xml:lang": "en",
    })
    martif.set("xmlns", "urn:iso:std:iso:30042:ed-1")

    # Header
    header = ET.SubElement(martif, "martifHeader")
    file_desc = ET.SubElement(header, "fileDesc")
    ET.SubElement(file_desc, "title").text = db.name or "TermBase"
    ET.SubElement(file_desc, "source").text = "TermPrep v0.3"
    encoding_desc = ET.SubElement(header, "encodingDesc")
    ET.SubElement(encoding_desc, "encoding").text = "UTF-8"

    # Body
    text = ET.SubElement(martif, "text")
    body = ET.SubElement(text, "body")

    for row in rows:
        word = row.get("word", "")
        translation = row.get("translation", "")
        if not word:
            continue

        term_entry = ET.SubElement(body, "termEntry", attrib={"id": f"t{row.get('id', '0')}"})

        # Source language term
        lang_set_src = ET.SubElement(term_entry, "langSet", attrib={
            "xml:lang": row.get("lang", "auto") or "en"
        })
        tig_src = ET.SubElement(lang_set_src, "tig")
        ET.SubElement(tig_src, "term").text = word
        if row.get("type"):
            ET.SubElement(tig_src, "termNote", attrib={"type": "partOfSpeech"}).text = row["type"]
        if row.get("domain"):
            ET.SubElement(tig_src, "termNote", attrib={"type": "subjectField"}).text = row["domain"]
        if row.get("status"):
            ET.SubElement(tig_src, "termNote", attrib={"type": "status"}).text = row["status"]

        # Target language term (if translation exists)
        if translation:
            lang_set_tgt = ET.SubElement(term_entry, "langSet", attrib={
                "xml:lang": "zh" if row.get("lang", "auto") in ("en", "auto") else "en"
            })
            tig_tgt = ET.SubElement(lang_set_tgt, "tig")
            ET.SubElement(tig_tgt, "term").text = translation

    # Pretty-print
    xml_bytes = ET.tostring(martif, encoding="UTF-8", xml_declaration=True)
    # Re-parse and write with indentation
    dom = ET.fromstring(xml_bytes)
    _indent_xml(dom)
    tree = ET.ElementTree(dom)
    tree.write(output_path, encoding="UTF-8", xml_declaration=True)
    return len(rows)


def export_json(db: TermDB, output_path: str, status_filter: str | None = None) -> int:
    """Export to JSON format for programmatic use."""
    rows = _fetch_rows(db, status_filter)
    stats = db.get_stats()
    data = {
        "termbase": db.name,
        "domain": db.domain,
        "total": len(rows),
        "stats": {
            "confirmed": stats.get("confirmed", 0),
            "by_domain": stats.get("by_domain", {}),
            "by_status": stats.get("by_status", {}),
        },
        "terms": rows,
    }
    with open(output_path, "w", encoding="UTF-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    return len(rows)


# ---- Internal helpers ----

def _fetch_rows(db: TermDB, status_filter: str | None = None) -> list[dict[str, Any]]:
    """Fetch terms from DB, optionally filtered by status."""
    with db._get_conn() as conn:
        if status_filter:
            rows = conn.execute(
                """SELECT id, word, translation, type, domain, status, lang, notes
                   FROM terms WHERE status = ? ORDER BY word""",
                (status_filter,)
            ).fetchall()
        else:
            rows = conn.execute(
                """SELECT id, word, translation, type, domain, status, lang, notes
                   FROM terms ORDER BY word"""
            ).fetchall()
    return [dict(r) for r in rows]


def _indent_xml(elem: ET.Element, level: int = 0) -> None:
    """Add indentation to XML element tree for pretty printing."""
    indent = "\n" + level * "  "
    if len(elem):
        if not elem.text or not elem.text.strip():
            elem.text = indent + "  "
        if not elem.tail or not elem.tail.strip():
            elem.tail = indent
        for child in elem:
            _indent_xml(child, level + 1)
        if not child.tail or not child.tail.strip():
            child.tail = indent
    else:
        if level and (not elem.tail or not elem.tail.strip()):
            elem.tail = indent


def _guess_lang(word: str) -> str:
    """Simple heuristic to guess if a word is Chinese or English."""
    import re
    if re.search(r"[\u4e00-\u9fff]", word):
        return "zh"
    return "en"
